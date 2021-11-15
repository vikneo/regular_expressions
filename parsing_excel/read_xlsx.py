import openpyxl

''' открываем книгу '''

book = openpyxl.open('name.xlsx',
                     read_only=True,  # атрибут только на чтение
                     data_only=True  # атрибут игнорирует все формулы
                     )
sheet = book.active  # метод для чтения первой и единственной страницы

''' 1-й вариант с помощью диапазона range() '''

for row in range(1, sheet.max_row + 1):
    col_1 = sheet[row][0].value
    col_2 = sheet[row][1].value
    col_3 = sheet[row][2].value
    col_4 = sheet[row][3].value
    print(row, col_1, col_2, col_3, col_4)

# =============================================================
''' 2-й вариант с выбором диапазона по колонкам '''

cells = sheet['B1':'D11']  # пример выбора колонок с какой по какую
for cell_1, cell_2 in cells:
    print(cell_1.value, cell_2.value)

# ==============================================================
''' 3-й вариант с помощью встроенного метода iter_rows '''

for row in sheet.iter_rows(min_row=5,   # начало с данного индекса строки
                           max_row=20,  # окончание на данном индексе строки
                           min_col=1,   # начало с данного индекса столбца
                           max_col=3    # окончание на данном индексе столбца
                           ):
    for cell in row:
        print(cell.value, end=' ')
    print()

'''  или возможно полное чтение страницы буз параметров '''
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value, end=' ')
    print()

''' Для получения полного списка страниц в книге '''

sheets = book.worksheets  # возможно обращение по индексу к странице  book.worksheets[1]
